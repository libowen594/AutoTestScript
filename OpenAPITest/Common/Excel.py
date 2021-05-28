#!/usr/bin/python
# -*- coding: UTF-8 -*-
__author__ = 'Bowen.li'

from openpyxl.utils import get_column_letter, column_index_from_string
from Common.ReadConfig import get_config_values
import os
import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles.colors import BLACK


class Excel:
    """文件写入数据"""

    def __init__(self):
        self._filename = "report" + str(datetime.datetime.today().date()) + ".xlsx"
        self._path = get_config_values("Report", "path")
        if not os.path.exists(os.path.join(self._path, self._filename)):
            book = Workbook()
            book.save(os.path.join(self._path, self._filename))
        self.__wb = None
        self.__ws = None

    def __OpenExcel(self, sheet_name=None):
        self.__wb = load_workbook(os.path.join(self._path, self._filename))
        if sheet_name:
            if sheet_name in self.__wb.sheetnames:
                self.__ws = self.__wb[sheet_name]
            else:
                print(f"将要打开的表{sheet_name}不存在,请创建后重新打开")
                exit()
        else:
            self.__ws = self.__wb.active

    def creat_sheet(self, sheet_name, header=None):
        """
        :param sheet_name:创建sheet的名称
        :param header: 创建表的表头
        :return:
        """
        self.__wb = load_workbook(os.path.join(self._path, self._filename))
        if sheet_name not in self.__wb.sheetnames:
            self.__wb.create_sheet(sheet_name, 0)
            if header:
                self.__ws = self.__wb[sheet_name]
                if isinstance(header, list):
                    self.__ws.append(header)
                    self.close()
                else:
                    self.close()
                    raise TypeError
            else:
                self.close()

        else:
            self.close()
            print(f"要创建的表:{sheet_name}已经存在")

    def write_one_cell(self, row: int, col: (int, str), value, sheet_name=None):
        """
        写入测试结果
        :param sheet_name: 需要写入数据的表的名称
        :param row: 需要写入单元格的行
        :param col: 需要写入单元格的列号
        :param value: 写入的值，以行为单位
        :return: 无
        """
        self.__OpenExcel(sheet_name)
        if not isinstance(row, int):
            raise TypeError
        if not isinstance(col, int):
            if col.isalpha():
                col = column_index_from_string(col.upper())
            else:
                raise TypeError
        self.__ws.cell(row=row, column=col, value=str(value))
        self.close()

    def write_rows(self, row: int, start_col: (int, str), end_col: (int, str), values: (list, str, int, dict),
                   sheet_name=None):
        """
        写入一行数据
        :param sheet_name: 需要写入的表的名称
        :param row: 需要写入的行号
        :param start_col: 需要写入的开始列号
        :param end_col: 需要写入的结束列号
        :param values: 需要写入的value列表
        :return:
        """
        self.__OpenExcel(sheet_name)
        if not isinstance(values, (list, str, int, dict)):
            raise TypeError
        if not isinstance(row, int):
            raise TypeError
        if not isinstance(start_col, int):
            if start_col.isalpha():
                start_col = column_index_from_string(start_col.upper())
            else:
                raise TypeError
        if not isinstance(end_col, int):
            if end_col.isalpha():
                end_col = column_index_from_string(end_col.upper())
            else:
                raise TypeError
        index = 0
        for cells in self.__ws.iter_rows(min_row=row, max_row=row, min_col=start_col, max_col=end_col):
            for cell in cells:
                if isinstance(values, list):
                    try:
                        cell.value = str(values[index])
                    except Exception:
                        cell.value = ""
                if isinstance(values, (str, int)):
                    cell.value = values
                index += 1
        self.close()

    def write_cols(self, col: (int, str), start_row: int, end_row: int, values: (list, str, int, dict),
                   sheet_name=None):
        """
        写入一列数据
        :param sheet_name: 需要写入表的名称
        :param col:需要写入列的列号，int或者字母
        :param start_row: 写入数据的开始行，行号
        :param end_row: 写入数据的结束行，行号
        :param values: 写入的value列表
        :return:
        """
        self.__OpenExcel(sheet_name)
        if not isinstance(values, (list, str, int, dict)):
            raise TypeError
        if not isinstance(start_row, int):
            raise TypeError
        if not isinstance(end_row, int):
            raise TypeError
        if not isinstance(col, int):
            if col.isalpha():
                col = column_index_from_string(col.upper())
            else:
                raise TypeError
        index = 0
        for cells in self.__ws.iter_cols(min_col=col, max_col=col, min_row=start_row, max_row=end_row):
            for cell in cells:
                if isinstance(values, list):
                    cell.value = str(values[index])
                if isinstance(values, (str, int)):
                    cell.value = values
                index += 1
        self.close()

    def get_value(self, row: int, col: (int, str), sheet_name=None):
        """
        获取一个单元格的值
        :param sheet_name:
        :param row: 行号，必须为int类型
        :param col: 列号，可以为int或者列的字母
        :return:
        """
        self.__OpenExcel(sheet_name)
        if not isinstance(row, int):
            raise TypeError
        if not isinstance(col, int):
            if col.isalpha():
                col = column_index_from_string(col.upper())
            else:
                raise TypeError
        for i in self.__ws.merged_cells:
            if i.min_row <= row <= i.max_row:
                row = i.min_row
            if i.min_col <= col <= i.max_col:
                col = i.min_col
        value = self.__ws.cell(row=row, column=col).value
        self.close()
        return value

    def get_row_num(self, value: str, sheet_name=None):
        """
        根据value获取这value在表中的行号
        :param sheet_name:
        :param value: 需要获取列号的值
        :return: 所有匹配行的集合
        """
        self.__OpenExcel(sheet_name)
        if not isinstance(value, str):
            raise TypeError
        row_list = []
        for row in self.__ws.iter_rows(min_row=1):
            for cell in row:
                if cell.value == value:
                    row_list.append(cell.row)
        self.close()
        if len(row_list) == 1:
            return row_list[0]
        else:
            return row_list

    def get_col_num(self, value: str, sheet_name=None):
        """
        根据value获取这value在表中的列号
        :param sheet_name:
        :param value: 需要获取列号的值
        :return: 所有匹配列的集合
        """
        self.__OpenExcel(sheet_name)
        if not isinstance(value, str):
            raise TypeError
        col_list = []
        for col in self.__ws.iter_cols(min_col=1):
            for cell in col:
                if cell.value == value:
                    col_list.append(cell.column)
        self.close()
        if len(col_list) == 1:
            return col_list[0]
        else:
            return col_list

    def get_rows_value(self, row_num: int, sheet_name=None):
        """
        获取一行数据
        :param sheet_name:
        :param row_num: 行号
        :return:
        """
        self.__OpenExcel(sheet_name)
        if not isinstance(row_num, int):
            raise TypeError
        for values in self.__ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True):
            return values
        self.close()

    def get_cols_value(self, col: (int, str), sheet_name=None):
        """
        获取一列数据
        :param sheet_name:
        :param col: 列号
        :return:
        """
        self.__OpenExcel(sheet_name)
        if not isinstance(col, int):
            if col.isalpha():
                col = column_index_from_string(col)
            else:
                raise TypeError
        for values in self.__ws.iter_cols(min_col=col, max_col=col, min_row=2, values_only=True):
            return values
        self.close()

    def get_sheet_values(self, sheet_name):
        self.__OpenExcel(sheet_name)
        data = []
        max_row = self.__ws.max_row
        for header in self.__ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for value in self.__ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
                data.append(dict(zip(header, value)))
        self.close()
        return data

    def get_sheets_name(self):
        self.__OpenExcel()
        return self.__wb.sheetnames

    def get_max_row(self, sheet_name=None):
        """
        获取表格中最大行的行号
        :return:
        """
        self.__OpenExcel(sheet_name)
        max_row = self.__ws.max_row
        self.close()
        return max_row

    def get_max_col(self, sheet_name=None):
        """
        获取表格中最大行的列号
        :return:
        """
        self.__OpenExcel(sheet_name)
        max_col = self.__ws.max_column
        self.close()
        return max_col

    def close(self):
        """
        关闭文档，并保存
        :return:
        """
        self._set_style()
        self.__wb.save(os.path.join(self._path, self._filename))

    def merge_cells(self, start_row_num: int, end_row_num: int, start_col_num: int, end_col_num: int):
        """
        合并单元格
        :param start_row_num:开始行号
        :param end_row_num:结束行号
        :param start_col_num:开始列号
        :param end_col_num:结束列号
        :return:
        """
        self.__ws.merge_cells(start_row=start_row_num, end_row=end_row_num, start_column=start_col_num,
                              end_column=end_col_num)
        self.close()

    def copy_sheet(self, src_sheet, tag_sheet):
        self.creat_sheet(tag_sheet)
        self.__OpenExcel(src_sheet)
        tag_ws = self.__wb[tag_sheet]
        for i, row in enumerate(self.__ws.iter_rows()):
            for j, cell in enumerate(row):
                tag_ws.cell(row=i + 1, column=j + 1, value=cell.value)
        self.__ws = tag_ws
        self.close()

    def _set_style(self):
        """
        设置表格属性
        :return:
        """
        try:
            cells = self.__ws["A1:" + str(get_column_letter(self.__ws.max_column)) + str(self.__ws.max_row)]
            align = Alignment(horizontal='center', vertical='center')
            border = Border(left=Side(border_style='thin',
                                      color=BLACK),
                            right=Side(border_style='thin',
                                       color=BLACK),
                            top=Side(border_style='thin',
                                     color=BLACK),
                            bottom=Side(border_style='thin',
                                        ))
            for i in cells:
                for j in i:
                    j.alignment = align
                    j.border = border
        except AttributeError:
            pass
